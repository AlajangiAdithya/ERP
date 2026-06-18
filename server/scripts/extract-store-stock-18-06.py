"""Extract Sheet2 of the 18-06-2026 Store Stock Statement into normalized JSON.

Sections are detected dynamically (the statement's row layout shifts between
revisions, so we anchor on the section-header keywords rather than fixed rows).
Output mirrors the field semantics the JS importer expects:
  section, category, name, qty, uom, batch, dom, doe, referredUnit, remarks

FIM (customer property, Sheet1) is intentionally excluded — it belongs to the
gate-pass flow, not regular product stock.

  python scripts/extract-store-stock-18-06.py <excel_path> <out_json>
"""
import json
import sys
from openpyxl import load_workbook

KEYWORDS = [
    ("FABRIC", "FABRICS", "Raw Material"),
    ("SPOOL", "SPOOLS", "Raw Material"),
    ("RESIN", "RESINS_HARDENERS", "Raw Material"),
    ("RUBBER", "RUBBER", "Raw Material"),
    ("SOLVENT", "SOLVENTS", "Raw Material"),
    ("CONSUMABLE", "CONSUMABLES", "Consumable"),
    ("STATIONERY", "STATIONERY", "Stationery"),
]


def cell(row, i):
    return row[i] if i < len(row) and row[i] is not None else None


def s(v):
    return str(v).strip() if v is not None else ""


# Placeholder tokens used in the sheet to mean "nothing here". Kept verbatim for
# the description/SL columns, but treated as null for batch/uom/ref/remarks so we
# don't create junk zero-qty batches with batchNo "-" or notes like "ref: -".
_PLACEHOLDERS = {"", "-", "--", "---", "na", "n/a", "nil", "none"}


def clean(v):
    t = s(v)
    return None if t.lower() in _PLACEHOLDERS else t


def parse_qty(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip()
    if t in ("", "-", "--", "NA", "na", "N/A"):
        return None
    try:
        return float(t)
    except ValueError:
        return None


def parse_date(v):
    # openpyxl returns datetime for real date cells; everything else (NA, "-",
    # multi-date strings) is left for the batch/remarks columns, date -> null.
    import datetime
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    return None


def pick_stock_sheet(wb):
    best, best_hits = wb.sheetnames[0], -1
    for sn in wb.sheetnames:
        ws = wb[sn]
        hits = 0
        for row in ws.iter_rows(values_only=True, max_row=300):
            c0, c1 = s(cell(row, 0)), s(cell(row, 1))
            if c0 and not c1 and any(k in c0.upper() for k, _, _ in KEYWORDS):
                hits += 1
        if hits > best_hits:
            best, best_hits = sn, hits
    return best


def main():
    excel_path = sys.argv[1]
    out_path = sys.argv[2]
    wb = load_workbook(excel_path, data_only=True)
    sheet = pick_stock_sheet(wb)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))

    # Detect section header rows in document order.
    headers = []  # (row_index, label, key, category)
    for i, r in enumerate(rows):
        c0, c1 = s(cell(r, 0)), s(cell(r, 1))
        if c0 and not c1:
            up = c0.upper()
            for kw, key, cat in KEYWORDS:
                if kw in up:
                    headers.append((i, c0, key, cat))
                    break

    out = []
    consumable_n = 0
    for hi, (idx, label, key, cat) in enumerate(headers):
        end = headers[hi + 1][0] if hi + 1 < len(headers) else len(rows)
        section_key = key
        if key == "CONSUMABLES":
            consumable_n += 1
            section_key = f"CONSUMABLES_{consumable_n}"
        for j in range(idx + 1, end):
            r = rows[j]
            c0, c1 = s(cell(r, 0)), s(cell(r, 1))
            if not c1:
                continue
            if c1.upper() == "ITEM DESCRIPTION" or c0.upper() == "SL NO":
                continue
            out.append({
                "section": section_key,
                "category": cat,
                "name": c1,
                "qty": parse_qty(cell(r, 2)),
                "uom": clean(cell(r, 3)),
                "batch": clean(cell(r, 4)),
                "dom": parse_date(cell(r, 5)),
                "doe": parse_date(cell(r, 6)),
                "referredUnit": clean(cell(r, 7)),
                "remarks": clean(cell(r, 8)),
            })
    wb.close()

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # Summary to stderr so stdout stays clean if piped.
    by_section = {}
    qty_pos = 0
    for o in out:
        by_section[o["section"]] = by_section.get(o["section"], 0) + 1
        if o["qty"] and o["qty"] > 0:
            qty_pos += 1
    print(f"sheet={sheet}  rows={len(out)}  with_qty>0={qty_pos}")
    for k, v in by_section.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
