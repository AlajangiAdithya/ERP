#!/usr/bin/env python3
"""
Extract Work Order data from the "FINAL LIST TO UPLOAD IN ERP" workbook into a
clean JSON file that import-work-orders.js loads into the ERP database.

Source : work-orders-final.xlsx  (override with WORKORDERS_XLSX), sheet
         "FINAL LIST TO UPLOAD IN ERP"
Output : work-orders-import.json (next to this script)

Sheet columns (row 2 is the header, data starts row 3):
  A Party Name   -> customerName
  B Order No     -> supplyOrderNo (rows sharing one Order No = one Work Order)
  C Date         -> supplyOrderDate
  D Description  -> WorkOrderItem.description (one item per row)
  E QTY          -> WorkOrderItem.quantity
  F Qty Billed   -> seeds deliveredQty / invoicedQty (in-progress)
  G Balance Qty  -> (informational; not stored)
  H PDC          -> pdcDate when it's a real date; otherwise the relative term
                    (e.g. "FIM+3 Months", "TO+45 Days") is kept in deliveryClause
                    and pdcDate is left null for Planning to fill in later.
  I Unit         -> assignedUnitName (raw text; importer resolves/creates it)
  J REMARKS      -> appended to remarks

Rules (agreed with the user):
  - Group rows by Order No -> one Work Order with multiple line items.
  - Keep EVERY order (none are dropped). Orders whose PDC is a relative term or
    blank import with pdcDate = null and the term saved in deliveryClause.
  - The raw Unit cell text is kept; the importer resolves it against the Unit
    records (creating the ones that don't exist yet). An order whose rows span
    several units is assigned to the FIRST unit, with the others noted in remarks.
  - Import as IN_PROGRESS, seeding billed qty so existing progress is kept.
  - The junk "Total"/blank-Order-No row is skipped.
"""
import openpyxl
import re
import json
import datetime
import os
from collections import OrderedDict

WORKBOOK = os.environ.get(
    "WORKORDERS_XLSX",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "work-orders-final.xlsx"),
)
SHEET = "FINAL LIST TO UPLOAD IN ERP"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "work-orders-import.json")

IMPORT_MARKER = "Imported from Cash Flow sheet (FINAL LIST)."


def s(v):
    if v is None:
        return None
    t = str(v).strip()
    return t or None


def num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        m = re.search(r"[-+]?\d*\.?\d+", str(v))
        return float(m.group()) if m else 0.0


def is_date(v):
    return isinstance(v, (datetime.datetime, datetime.date))


def unit_name(v):
    """Clean an Excel 'Unit' cell; the importer resolves/creates it."""
    t = s(v)
    if not t or t.upper() in ("NA", "N/A", "-"):
        return None
    return re.sub(r"\s+", " ", t)


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]

    def cell(r, c):
        return ws.cell(row=r, column=c).value

    # Group rows by Order No (col B). Rows without an Order No are skipped
    # (the trailing "Total" summary row).
    groups = OrderedDict()
    for r in range(3, ws.max_row + 1):
        order_no = s(cell(r, 2))
        if not order_no:
            continue
        groups.setdefault(order_no, []).append(r)

    orders = []
    no_pdc = 0
    for order_no, rows in groups.items():
        # PDC: first real date wins; else keep the relative term as a clause.
        pdc = next((cell(r, 8) for r in rows if is_date(cell(r, 8))), None)
        delivery_clause = None
        if pdc is None:
            term = next((s(cell(r, 8)) for r in rows if s(cell(r, 8))), None)
            if term:
                delivery_clause = f"PDC: {term}"
            no_pdc += 1

        # Order date: first real date in col C, else fall back to the PDC date.
        odate = next((cell(r, 3) for r in rows if is_date(cell(r, 3))), None) or pdc
        party = next((s(cell(r, 1)) for r in rows if s(cell(r, 1))), None)

        # Distinct units in sheet order. First = primary assignment; the rest are
        # recorded in remarks so a multi-unit order still shows where else it ran.
        unit_names = []
        for r in rows:
            n = unit_name(cell(r, 9))
            if n and n not in unit_names:
                unit_names.append(n)
        assigned = unit_names[0] if unit_names else None
        other_units = unit_names[1:]

        items, qty_total, billed_total = [], 0.0, 0.0
        for i, r in enumerate(rows, 1):
            q = num(cell(r, 5))
            items.append({
                "lineNo": i,
                "description": s(cell(r, 4)) or "Material",
                "quantity": q,
                "uom": "Nos",
            })
            qty_total += q
            billed_total += num(cell(r, 6))
        billed_total = min(billed_total, qty_total)

        remark_parts = []
        sheet_remarks = [s(cell(r, 10)) for r in rows if s(cell(r, 10))]
        # De-dup while preserving order.
        seen = set()
        for rm in sheet_remarks:
            if rm not in seen:
                seen.add(rm)
                remark_parts.append(rm)
        if other_units:
            remark_parts.append("Also assigned: " + " / ".join(other_units))
        remark_parts.append(IMPORT_MARKER)

        orders.append({
            "supplyOrderNo": order_no,
            "supplyOrderDate": odate.date().isoformat() if is_date(odate) else None,
            "pdcDate": pdc.date().isoformat() if is_date(pdc) else None,
            "deliveryClause": delivery_clause,
            "customerName": party or "Not Provided",
            "nomenclature": (items[0]["description"][:120] if items else None),
            "assignedUnitName": assigned,
            "orderQuantity": round(qty_total, 4),
            "billedQty": round(billed_total, 4),
            "deliveryStatus": "PARTIAL" if billed_total > 0 else "IN_PROGRESS",
            "status": "IN_PROGRESS",
            "remarks": " | ".join(remark_parts),
            "items": items,
        })

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(orders, f, indent=2, ensure_ascii=False)

    named_n = sum(1 for o in orders if o["assignedUnitName"])
    print("Orders to import      :", len(orders))
    print("Line items            :", sum(len(o["items"]) for o in orders))
    print("Unit name present     :", named_n)
    print("Unit name blank       :", len(orders) - named_n)
    print("PDC empty (relative)  :", no_pdc)
    print("Distinct unit names   :",
          sorted({o["assignedUnitName"] for o in orders if o["assignedUnitName"]}))
    print("Written               :", OUT)


if __name__ == "__main__":
    main()
